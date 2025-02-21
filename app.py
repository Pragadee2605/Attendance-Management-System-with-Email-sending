from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, flash
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from flask import session

app = Flask(__name__)
app.secret_key = "your_secret_key"  # Required for flash messages

SENDER_EMAIL = " "
SENDER_PASSWORD = " "

DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD = "password123"

EXCEL_FILE = 'attendance1.xlsx'

# Initialize Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Reg Number", "Student Name", "Attendance (%)", "Status"])  # Header
    wb.save(EXCEL_FILE)


@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == DEFAULT_USERNAME and password == DEFAULT_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            flash("Invalid Username or Password", "danger")
    
    return render_template('login.html')




@app.route('/index', methods=['GET', 'POST'])
def index():
    date_filter = request.args.get('date')  # Retrieve the selected date from the request
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    attendance_data = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[2] and row[3] is not None:
            if date_filter:  # Filter by date if provided
                if row[0] == date_filter:
                    attendance_data.append({'name': row[2], 'percentage': row[3]})
            else:
                attendance_data.append({'name': row[2], 'percentage': row[3]})

    return render_template('index.html', attendance_data=attendance_data, selected_date=date_filter)

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        reg_number = request.form['reg_number']
        name = request.form['name']
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Check if the student already exists by reg_number
        for row in ws.iter_rows(min_row=2):  # Skip header
            if row[1].value == reg_number:
                flash("Student with this Reg Number already exists!", "danger")
                return redirect(url_for('add_student'))

        ws.append([datetime.now().strftime("%Y-%m-%d"), reg_number, name, 100, "Present"])
        wb.save(EXCEL_FILE)
        return redirect(url_for('view_students'))
    return render_template('add_student.html')

@app.route('/view_students')
def view_students():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        students.append(row)
    return render_template('view_students.html', students=students)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    reg_number = request.form['reg_number']
    status = request.form['status']
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if row[1].value == reg_number:  # Match Reg Number
            if status == 'Absent':
                row[3].value = max(0, row[3].value - 5)  # Decrease attendance %, ensure it doesn't go below 0
            else:
                row[3].value = min(100, row[3].value + 5)  # Increase attendance %, ensure it doesn't exceed 100
            row[4].value = status  # Update status
            row[0].value = datetime.now().strftime("%Y-%m-%d")  # Update date
            break  # Exit loop once the student is found
    
    wb.save(EXCEL_FILE)
    return redirect(url_for('view_students'))

@app.route('/filter_by_date', methods=['GET', 'POST'])
def filter_by_date():
    if request.method == 'POST':
        date = request.form['date']
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        filtered_records = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0] == date]
        return render_template('view_students.html', students=filtered_records)
    return render_template('filter_by_date.html')

@app.route('/filter_below_75', methods=['GET', 'POST'])
def filter_below_75():
    if request.method == 'POST':
        date = request.form['date']
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        filtered_students = [
            row for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0] == date and row[3] <= 75
        ]
        return render_template('view_students.html', students=filtered_students)
    return render_template('filter_below_75.html')

@app.route('/download/<date>')
def download(date):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    filtered_records = [ws[1]]  # Include header
    for row in ws.iter_rows(min_row=2):
        if row[0].value == date:
            filtered_records.append([cell.value for cell in row])

    download_file = f'attendance_{date}.xlsx'
    wb_filtered = Workbook()
    ws_filtered = wb_filtered.active
    for record in filtered_records:
        ws_filtered.append(record)
    wb_filtered.save(download_file)

    return send_file(download_file, as_attachment=True)

@app.route('/get_attendance', methods=['POST'])
def get_attendance():
    reg_number = request.json.get('reg_number')
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    attendance_details = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == reg_number:
            attendance_details.append({
                'date': row[0],
                'name': row[2],
                'percentage': row[3],
                'status': row[4]
            })

    return jsonify(attendance_details)


# Function to send email
@app.route("/message", methods=["GET", "POST"])
def send_email():
    if request.method == "POST":
        recipient_email = request.form.get("recipient_email")  # Get email from form
        
        if recipient_email:  # Check if the email is provided
            try:
                msg = MIMEText("This email is to notify you that your attendance is less than 75% in your academics. Kindly check it yourself!")
                msg["Subject"] = "Test Email"
                msg["From"] = SENDER_EMAIL
                msg["To"] = recipient_email

                # Connect to Gmail SMTP server
                with smtplib.SMTP("smtp.gmail.com", 587) as server:
                    server.starttls()  # Secure connection
                    server.login(SENDER_EMAIL, SENDER_PASSWORD)
                    server.sendmail(SENDER_EMAIL, recipient_email, msg.as_string())

                flash("Email sent successfully!", "success")
            except Exception as e:
                flash(f"Error sending email: {e}", "danger")
        else:
            flash("Recipient email is required!", "danger")

    return render_template("message.html")  # You need to create the message.html template with a form for email


if __name__ == '__main__':
    app.run(debug=True, port=5001)
